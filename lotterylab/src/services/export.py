"""
Export service for generating PDF and Excel reports.

This module provides functionality to export lottery analysis results
to various formats for offline viewing and sharing.
"""

from __future__ import annotations
from typing import Dict, Any, Optional, List
from datetime import datetime
from io import BytesIO
import pandas as pd

# PDF generation
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, 
    PageBreak, Image
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportExporter:
    """
    Generates PDF and Excel reports from lottery analysis data.
    """
    
    # Branding colors
    PRIMARY_COLOR = HexColor("#003366")
    ACCENT_COLOR = HexColor("#0066CC")
    SUCCESS_COLOR = HexColor("#00835C")
    WARNING_COLOR = HexColor("#B25E00")
    
    def __init__(self, language: str = "en"):
        self.language = language
        self.texts = self._get_texts()
    
    def _get_texts(self) -> Dict[str, str]:
        """Get localized text strings."""
        if self.language == "pl":
            return {
                "report_title": "Raport Analizy Loterii",
                "generated_on": "Wygenerowano",
                "game_type": "Typ gry",
                "analysis_period": "Okres analizy",
                "days": "dni",
                "frequency_analysis": "Analiza Częstości",
                "number": "Liczba",
                "frequency": "Częstość",
                "expected": "Oczekiwana",
                "delta": "Delta",
                "hot_numbers": "Gorące Liczby",
                "cold_numbers": "Zimne Liczby",
                "randomness_analysis": "Analiza Losowości",
                "test_name": "Nazwa testu",
                "statistic": "Statystyka",
                "p_value": "Wartość p",
                "result": "Wynik",
                "random": "Losowy",
                "not_random": "Nie losowy",
                "pattern_analysis": "Analiza Wzorców",
                "total_draws": "Liczba losowań",
                "disclaimer": "Ten raport służy wyłącznie celom edukacyjnym i rozrywkowym. Wyniki loterii są nieprzewidywalne.",
                "page": "Strona",
            }
        return {
            "report_title": "Lottery Analysis Report",
            "generated_on": "Generated on",
            "game_type": "Game type",
            "analysis_period": "Analysis period",
            "days": "days",
            "frequency_analysis": "Frequency Analysis",
            "number": "Number",
            "frequency": "Frequency",
            "expected": "Expected",
            "delta": "Delta",
            "hot_numbers": "Hot Numbers",
            "cold_numbers": "Cold Numbers",
            "randomness_analysis": "Randomness Analysis",
            "test_name": "Test Name",
            "statistic": "Statistic",
            "p_value": "P-Value",
            "result": "Result",
            "random": "Random",
            "not_random": "Not Random",
            "pattern_analysis": "Pattern Analysis",
            "total_draws": "Total draws",
            "disclaimer": "This report is for educational and entertainment purposes only. Lottery results are unpredictable.",
            "page": "Page",
        }
    
    def generate_pdf(
        self,
        frequency_data: Optional[Dict] = None,
        randomness_data: Optional[Dict] = None,
        patterns_data: Optional[Dict] = None,
        game_type: str = "lotto",
        window_days: int = 365
    ) -> BytesIO:
        """
        Generate a PDF report with analysis results.
        
        Args:
            frequency_data: Frequency analysis results
            randomness_data: Randomness test results
            patterns_data: Pattern analysis results
            game_type: Type of lottery game
            window_days: Analysis window in days
            
        Returns:
            BytesIO buffer containing the PDF
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=self.PRIMARY_COLOR,
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=self.ACCENT_COLOR,
            spaceBefore=20,
            spaceAfter=10
        )
        
        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=8
        )
        
        elements = []
        
        # Title
        elements.append(Paragraph(self.texts["report_title"], title_style))
        elements.append(Spacer(1, 6*mm))
        
        # Metadata
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        meta_text = f"""
        <b>{self.texts['generated_on']}:</b> {now}<br/>
        <b>{self.texts['game_type']}:</b> {game_type.upper()}<br/>
        <b>{self.texts['analysis_period']}:</b> {window_days} {self.texts['days']}
        """
        elements.append(Paragraph(meta_text, body_style))
        elements.append(Spacer(1, 10*mm))
        
        # Frequency Analysis Section
        if frequency_data:
            elements.append(Paragraph(f"📊 {self.texts['frequency_analysis']}", section_style))
            
            # Summary
            num_draws = frequency_data.get('num_draws', 0)
            expected = frequency_data.get('expected_each', 0)
            elements.append(Paragraph(
                f"{self.texts['total_draws']}: {num_draws} | {self.texts['expected']}: {expected:.2f}",
                body_style
            ))
            
            # Hot/Cold numbers
            hot = frequency_data.get('hot_numbers', [])
            cold = frequency_data.get('cold_numbers', [])
            if hot:
                elements.append(Paragraph(
                    f"🔥 {self.texts['hot_numbers']}: {', '.join(map(str, hot))}",
                    body_style
                ))
            if cold:
                elements.append(Paragraph(
                    f"❄️ {self.texts['cold_numbers']}: {', '.join(map(str, cold))}",
                    body_style
                ))
            
            # Frequency table (top 10 most/least frequent)
            freq = frequency_data.get('frequency', {})
            if freq:
                sorted_freq = sorted(freq.items(), key=lambda x: x[1], reverse=True)
                table_data = [
                    [self.texts['number'], self.texts['frequency'], self.texts['delta']]
                ]
                
                # Top 5 + Bottom 5
                for num, count in sorted_freq[:5] + sorted_freq[-5:]:
                    delta = count - expected
                    table_data.append([str(num), str(count), f"{delta:+.1f}"])
                
                table = Table(table_data, colWidths=[40*mm, 40*mm, 40*mm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.ACCENT_COLOR),
                    ('TEXTCOLOR', (0, 0), (-1, 0), HexColor("#FFFFFF")),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor("#F5F5F5"), HexColor("#FFFFFF")]),
                ]))
                elements.append(Spacer(1, 5*mm))
                elements.append(table)
        
        # Randomness Analysis Section
        if randomness_data:
            elements.append(Spacer(1, 10*mm))
            elements.append(Paragraph(f"🎲 {self.texts['randomness_analysis']}", section_style))
            
            # Tests table
            tests = []
            
            if 'chi_square' in randomness_data:
                cs = randomness_data['chi_square']
                tests.append([
                    "Chi-Square",
                    f"{cs.get('chi_square_statistic', 0):.2f}",
                    f"{cs.get('p_value', 0):.4f}",
                    self.texts['random'] if cs.get('p_value', 0) > 0.05 else self.texts['not_random']
                ])
            
            if 'kolmogorov_smirnov' in randomness_data:
                ks = randomness_data['kolmogorov_smirnov']
                tests.append([
                    "Kolmogorov-Smirnov",
                    f"{ks.get('ks_statistic', 0):.4f}",
                    f"{ks.get('p_value', 0):.4f}",
                    self.texts['random'] if ks.get('p_value', 0) > 0.05 else self.texts['not_random']
                ])
            
            if 'runs_test' in randomness_data:
                runs = randomness_data['runs_test']
                tests.append([
                    "Runs Test",
                    f"{runs.get('runs_count', 0)}",
                    f"{runs.get('p_value', 0):.4f}",
                    self.texts['random'] if runs.get('p_value', 0) > 0.05 else self.texts['not_random']
                ])
            
            if tests:
                table_data = [
                    [self.texts['test_name'], self.texts['statistic'], 
                     self.texts['p_value'], self.texts['result']]
                ] + tests
                
                table = Table(table_data, colWidths=[45*mm, 35*mm, 35*mm, 35*mm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.ACCENT_COLOR),
                    ('TEXTCOLOR', (0, 0), (-1, 0), HexColor("#FFFFFF")),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor("#F5F5F5"), HexColor("#FFFFFF")]),
                ]))
                elements.append(Spacer(1, 5*mm))
                elements.append(table)
        
        # Pattern Analysis Section
        if patterns_data:
            elements.append(Spacer(1, 10*mm))
            elements.append(Paragraph(f"🔍 {self.texts['pattern_analysis']}", section_style))
            
            if 'consecutive' in patterns_data:
                cons = patterns_data['consecutive']
                elements.append(Paragraph(
                    f"• Consecutive: {cons.get('total_consecutive', 0)} sequences, "
                    f"max length: {cons.get('max_length', 0)}",
                    body_style
                ))
            
            if 'arithmetic' in patterns_data:
                arith = patterns_data['arithmetic']
                elements.append(Paragraph(
                    f"• Arithmetic sequences: {arith.get('total_sequences', 0)}",
                    body_style
                ))
            
            if 'sums' in patterns_data:
                sums = patterns_data['sums']
                elements.append(Paragraph(
                    f"• Sum range: {sums.get('min_sum', 0)} - {sums.get('max_sum', 0)}, "
                    f"average: {sums.get('avg_sum', 0):.1f}",
                    body_style
                ))
        
        # Disclaimer
        elements.append(Spacer(1, 20*mm))
        disclaimer_style = ParagraphStyle(
            'Disclaimer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=HexColor("#666666"),
            alignment=TA_CENTER,
            borderColor=HexColor("#CCCCCC"),
            borderWidth=1,
            borderPadding=10
        )
        elements.append(Paragraph(f"⚠️ {self.texts['disclaimer']}", disclaimer_style))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def generate_excel(
        self,
        frequency_data: Optional[Dict] = None,
        randomness_data: Optional[Dict] = None,
        patterns_data: Optional[Dict] = None,
        draws_data: Optional[List[Dict]] = None,
        game_type: str = "lotto",
        window_days: int = 365
    ) -> BytesIO:
        """
        Generate an Excel report with analysis results.
        
        Args:
            frequency_data: Frequency analysis results
            randomness_data: Randomness test results
            patterns_data: Pattern analysis results
            draws_data: Recent draws data
            game_type: Type of lottery game
            window_days: Analysis window in days
            
        Returns:
            BytesIO buffer containing the Excel file
        """
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        accent_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Remove default sheet
        default_sheet = wb.active
        
        # Summary Sheet
        ws_summary = wb.create_sheet(title="Summary")
        ws_summary['A1'] = self.texts['report_title']
        ws_summary['A1'].font = Font(bold=True, size=16, color="003366")
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A3'] = self.texts['generated_on']
        ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws_summary['A4'] = self.texts['game_type']
        ws_summary['B4'] = game_type.upper()
        ws_summary['A5'] = self.texts['analysis_period']
        ws_summary['B5'] = f"{window_days} {self.texts['days']}"
        
        if frequency_data:
            ws_summary['A7'] = self.texts['total_draws']
            ws_summary['B7'] = frequency_data.get('num_draws', 0)
            ws_summary['A8'] = self.texts['hot_numbers']
            ws_summary['B8'] = ', '.join(map(str, frequency_data.get('hot_numbers', [])))
            ws_summary['A9'] = self.texts['cold_numbers']
            ws_summary['B9'] = ', '.join(map(str, frequency_data.get('cold_numbers', [])))
        
        # Frequency Sheet
        if frequency_data and frequency_data.get('frequency'):
            ws_freq = wb.create_sheet(title="Frequency")
            
            headers = [self.texts['number'], self.texts['frequency'], 
                      self.texts['expected'], self.texts['delta'], "%"]
            for col, header in enumerate(headers, 1):
                cell = ws_freq.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            freq = frequency_data['frequency']
            expected = frequency_data.get('expected_each', 0)
            
            for row, (num, count) in enumerate(sorted(freq.items()), 2):
                delta = count - expected
                pct_delta = (delta / expected * 100) if expected > 0 else 0
                
                ws_freq.cell(row=row, column=1, value=num).border = thin_border
                ws_freq.cell(row=row, column=2, value=count).border = thin_border
                ws_freq.cell(row=row, column=3, value=round(expected, 2)).border = thin_border
                ws_freq.cell(row=row, column=4, value=round(delta, 2)).border = thin_border
                ws_freq.cell(row=row, column=5, value=round(pct_delta, 1)).border = thin_border
            
            # Auto-fit columns
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws_freq.column_dimensions[col].width = 12
        
        # Randomness Sheet
        if randomness_data:
            ws_rand = wb.create_sheet(title="Randomness")
            
            headers = [self.texts['test_name'], self.texts['statistic'],
                      self.texts['p_value'], self.texts['result']]
            for col, header in enumerate(headers, 1):
                cell = ws_rand.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            row = 2
            tests_to_export = [
                ('chi_square', 'Chi-Square', 'chi_square_statistic'),
                ('kolmogorov_smirnov', 'Kolmogorov-Smirnov', 'ks_statistic'),
                ('runs_test', 'Runs Test', 'runs_count'),
            ]
            
            for key, name, stat_key in tests_to_export:
                if key in randomness_data:
                    test = randomness_data[key]
                    p_val = test.get('p_value', 0)
                    
                    ws_rand.cell(row=row, column=1, value=name).border = thin_border
                    ws_rand.cell(row=row, column=2, value=round(test.get(stat_key, 0), 4)).border = thin_border
                    ws_rand.cell(row=row, column=3, value=round(p_val, 4)).border = thin_border
                    ws_rand.cell(row=row, column=4, 
                               value=self.texts['random'] if p_val > 0.05 else self.texts['not_random']).border = thin_border
                    row += 1
            
            for col in ['A', 'B', 'C', 'D']:
                ws_rand.column_dimensions[col].width = 20
        
        # Draws Sheet
        if draws_data:
            ws_draws = wb.create_sheet(title="Recent Draws")
            
            headers = ["#", self.texts.get('date', 'Date'), 
                      self.texts.get('numbers', 'Numbers'), self.texts.get('type', 'Type')]
            for col, header in enumerate(headers, 1):
                cell = ws_draws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            for row, draw in enumerate(draws_data[:100], 2):  # Max 100 draws
                ws_draws.cell(row=row, column=1, value=draw.get('draw_number', '')).border = thin_border
                ws_draws.cell(row=row, column=2, value=draw.get('draw_date', '')).border = thin_border
                ws_draws.cell(row=row, column=3, value=draw.get('numbers', '')).border = thin_border
                ws_draws.cell(row=row, column=4, value=draw.get('game_type', '')).border = thin_border
            
            ws_draws.column_dimensions['A'].width = 10
            ws_draws.column_dimensions['B'].width = 15
            ws_draws.column_dimensions['C'].width = 25
            ws_draws.column_dimensions['D'].width = 12
        
        # Remove the default empty sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


def create_frequency_report(
    frequency_data: Dict,
    format: str = "pdf",
    language: str = "en"
) -> BytesIO:
    """
    Convenience function to create a frequency-only report.
    """
    exporter = ReportExporter(language=language)
    
    if format == "excel":
        return exporter.generate_excel(
            frequency_data=frequency_data,
            game_type=frequency_data.get('game_type', 'lotto'),
            window_days=frequency_data.get('window_days', 365)
        )
    else:
        return exporter.generate_pdf(
            frequency_data=frequency_data,
            game_type=frequency_data.get('game_type', 'lotto'),
            window_days=frequency_data.get('window_days', 365)
        )


def create_full_report(
    frequency_data: Optional[Dict] = None,
    randomness_data: Optional[Dict] = None,
    patterns_data: Optional[Dict] = None,
    draws_data: Optional[List[Dict]] = None,
    format: str = "pdf",
    language: str = "en",
    game_type: str = "lotto",
    window_days: int = 365
) -> BytesIO:
    """
    Convenience function to create a comprehensive report.
    """
    exporter = ReportExporter(language=language)
    
    if format == "excel":
        return exporter.generate_excel(
            frequency_data=frequency_data,
            randomness_data=randomness_data,
            patterns_data=patterns_data,
            draws_data=draws_data,
            game_type=game_type,
            window_days=window_days
        )
    else:
        return exporter.generate_pdf(
            frequency_data=frequency_data,
            randomness_data=randomness_data,
            patterns_data=patterns_data,
            game_type=game_type,
            window_days=window_days
        )


